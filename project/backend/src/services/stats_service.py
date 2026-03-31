"""
Агрегация метрик для статистики и отчётов (ошибки, локации, время решения, исполнители).
"""
from __future__ import annotations

import io
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple
import uuid

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func
from sqlalchemy.orm import Session

from src.core.models import ErrorGroup, ErrorTask, Event, EventError, EventUrl


def parse_iso_utc(s: Optional[str]) -> Optional[datetime]:
    if not s or not str(s).strip():
        return None
    raw = str(s).strip().replace("Z", "+00:00")
    dt = datetime.fromisoformat(raw)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def default_range_days(days: int = 7) -> Tuple[datetime, datetime]:
    end = datetime.now(timezone.utc)
    start = end - timedelta(days=days)
    return start, end


def resolve_time_range(
    date_from: Optional[str],
    date_to: Optional[str],
    default_days: int = 7,
) -> Tuple[datetime, datetime]:
    end = parse_iso_utc(date_to) if date_to else datetime.now(timezone.utc)
    start = parse_iso_utc(date_from) if date_from else None
    if start is None:
        start = end - timedelta(days=default_days)
    if end < start:
        start, end = end, start
    return start, end


def aggregate_metrics(
    db: Session,
    start: datetime,
    end: datetime,
    project_id: Optional[uuid.UUID] = None,
) -> Dict[str, Any]:
    """Сводные метрики за [start, end)."""
    ev_base = (
        db.query(Event)
        .join(EventError, EventError.event_id == Event.id)
        .filter(Event.created_at >= start, Event.created_at < end)
    )
    if project_id:
        ev_base = ev_base.filter(Event.project_id == project_id)

    events_with_errors = ev_base.count()

    dom_col = func.coalesce(EventUrl.domain, "(нет домена)")
    domain_rows = (
        db.query(dom_col, func.count(Event.id))
        .select_from(Event)
        .join(EventError, EventError.event_id == Event.id)
        .outerjoin(EventUrl, EventUrl.event_id == Event.id)
        .filter(Event.created_at >= start, Event.created_at < end)
    )
    if project_id:
        domain_rows = domain_rows.filter(Event.project_id == project_id)
    domain_rows = (
        domain_rows.group_by(dom_col).order_by(func.count(Event.id).desc()).limit(20).all()
    )
    top_domains = [{"domain": r[0], "count": int(r[1])} for r in domain_rows]

    group_rows = (
        db.query(ErrorGroup.title, func.count(Event.id).label("cnt"))
        .join(Event, Event.error_group_id == ErrorGroup.id)
        .filter(
            Event.error_group_id.isnot(None),
            Event.created_at >= start,
            Event.created_at < end,
        )
    )
    if project_id:
        group_rows = group_rows.filter(Event.project_id == project_id)
    group_rows = group_rows.group_by(ErrorGroup.title).order_by(func.count(Event.id).desc()).limit(20).all()
    top_groups = [{"title": r[0], "count": int(r.cnt)} for r in group_rows]

    task_q = db.query(ErrorTask).filter(
        ErrorTask.created_at >= start,
        ErrorTask.created_at < end,
    )
    if project_id:
        task_q = task_q.filter(ErrorTask.project_id == project_id)
    tasks_created = task_q.count()

    resolved_in_period = db.query(ErrorTask).filter(
        ErrorTask.is_resolved.is_(True),
        ErrorTask.resolved_at.isnot(None),
        ErrorTask.resolved_at >= start,
        ErrorTask.resolved_at < end,
    )
    if project_id:
        resolved_in_period = resolved_in_period.filter(ErrorTask.project_id == project_id)
    resolved_count = resolved_in_period.count()

    sec_expr = func.extract("epoch", ErrorTask.resolved_at - ErrorTask.created_at)
    rstats = (
        db.query(
            func.min(sec_expr),
            func.max(sec_expr),
            func.avg(sec_expr),
        )
        .select_from(ErrorTask)
        .filter(
            ErrorTask.is_resolved.is_(True),
            ErrorTask.resolved_at.isnot(None),
            ErrorTask.resolved_at >= start,
            ErrorTask.resolved_at < end,
        )
    )
    if project_id:
        rstats = rstats.filter(ErrorTask.project_id == project_id)
    rmin, rmax, ravg = rstats.one()
    resolution_seconds = {
        "min": float(rmin) if rmin is not None else None,
        "max": float(rmax) if rmax is not None else None,
        "avg": float(ravg) if ravg is not None else None,
    }

    ack_lbl = func.coalesce(ErrorTask.acknowledged_by_label, "(неизвестно)")
    ack_rows = (
        db.query(ack_lbl, func.count(ErrorTask.id))
        .filter(
            ErrorTask.is_acknowledged.is_(True),
            ErrorTask.acknowledged_at.isnot(None),
            ErrorTask.created_at >= start,
            ErrorTask.created_at < end,
        )
    )
    if project_id:
        ack_rows = ack_rows.filter(ErrorTask.project_id == project_id)
    ack_rows = ack_rows.group_by(ack_lbl).order_by(func.count(ErrorTask.id).desc()).limit(15)
    top_took_task = [{"label": r[0], "count": int(r[1])} for r in ack_rows.all()]

    res_lbl = func.coalesce(ErrorTask.resolved_by_label, "(неизвестно)")
    res_rows = (
        db.query(res_lbl, func.count(ErrorTask.id))
        .filter(
            ErrorTask.is_resolved.is_(True),
            ErrorTask.resolved_at.isnot(None),
            ErrorTask.resolved_at >= start,
            ErrorTask.resolved_at < end,
        )
    )
    if project_id:
        res_rows = res_rows.filter(ErrorTask.project_id == project_id)
    res_rows = res_rows.group_by(res_lbl).order_by(func.count(ErrorTask.id).desc()).limit(15)
    top_resolved = [{"label": r[0], "count": int(r[1])} for r in res_rows.all()]

    return {
        "period": {
            "start": start.isoformat(),
            "end": end.isoformat(),
        },
        "project_id": str(project_id) if project_id else None,
        "events_with_errors": events_with_errors,
        "tasks_created": tasks_created,
        "tasks_resolved": resolved_count,
        "top_domains": top_domains,
        "top_error_groups": top_groups,
        "resolution_time_seconds": resolution_seconds,
        "top_who_took_task": top_took_task,
        "top_who_resolved": top_resolved,
    }


def build_excel_report(
    db: Session,
    start: datetime,
    end: datetime,
    project_id: Optional[uuid.UUID] = None,
    project_label: Optional[str] = None,
) -> bytes:
    data = aggregate_metrics(db, start, end, project_id)
    wb = Workbook()
    bold = Font(bold=True)

    project_cell = project_label or (data["project_id"] or "все")

    ws0 = wb.active
    ws0.title = "Сводка"
    ws0.append(["Период", f"{data['period']['start']} — {data['period']['end']}"])
    ws0.append(["Проект", project_cell])
    ws0.append(["Событий с ошибкой (Event+EventError)", data["events_with_errors"]])
    ws0.append(["Задач создано (ErrorTask)", data["tasks_created"]])
    ws0.append(["Задач решено", data["tasks_resolved"]])
    rts = data["resolution_time_seconds"]
    ws0.append([])
    ws0.append(["Время решения (сек)", ""])
    ws0.append(["min", rts["min"]])
    ws0.append(["max", rts["max"]])
    ws0.append(["avg", rts["avg"]])
    for row in ws0.iter_rows(min_row=1, max_row=1):
        for c in row:
            c.font = bold
    for row in ws0.iter_rows(min_row=7, max_row=7):
        for c in row:
            c.font = bold

    ws1 = wb.create_sheet("Домены")
    ws1.append(["Домен", "Событий"])
    ws1["A1"].font = bold
    ws1["B1"].font = bold
    for row in data["top_domains"]:
        ws1.append([row["domain"], row["count"]])

    ws2 = wb.create_sheet("Группы ошибок")
    ws2.append(["Группа (title)", "Событий"])
    ws2["A1"].font = bold
    ws2["B1"].font = bold
    for row in data["top_error_groups"]:
        ws2.append([row["title"], row["count"]])

    ws3 = wb.create_sheet("Исполнители")
    ws3.append(["Кто брал в работу", "Задач"])
    ws3["A1"].font = bold
    ws3["B1"].font = bold
    for row in data["top_who_took_task"]:
        ws3.append([row["label"], row["count"]])
    ws3.append([])
    ws3.append(["Кто закрыл", "Задач"])
    ws3.cell(row=ws3.max_row, column=1).font = bold
    ws3.cell(row=ws3.max_row, column=2).font = bold
    for row in data["top_who_resolved"]:
        ws3.append([row["label"], row["count"]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_excel_to_telegram(xlsx_bytes: bytes, caption: str, filename: str) -> None:
    """Отправка .xlsx в общий чат (TELEGRAM_CHAT_ID в .env), если задан."""
    import logging
    import requests

    from src.core.config import get_settings

    logger = logging.getLogger(__name__)
    s = get_settings()
    if not s.telegram_chat_id or not str(s.telegram_chat_id).strip():
        logger.warning("TELEGRAM_CHAT_ID не задан — отправка Excel отчёта пропущена")
        return
    url = f"https://api.telegram.org/bot{s.telegram_bot_token}/sendDocument"
    files = {"document": (filename, xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    tid = str(s.telegram_chat_id).strip().replace(" ", "")
    if tid.startswith("-") and tid[1:].isdigit():
        chat_id_val = int(tid)
    elif tid.isdigit():
        chat_id_val = int(tid)
    else:
        chat_id_val = tid
    data = {"chat_id": chat_id_val, "caption": caption[:1024]}
    resp = requests.post(url, files=files, data=data, timeout=120)
    resp.raise_for_status()
